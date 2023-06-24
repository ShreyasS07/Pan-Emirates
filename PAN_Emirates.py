import os
import time
import psutil
import warnings
import pandas as pd
import tkinter as tk
from openpyxl import Workbook
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from tkinter.messagebox import askyesno
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=FutureWarning)


def input_file():
    global excel_path
    excel_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx;*.xls;*.csv')])   # Select Excel File
    print("\nSelected Label File path:", excel_path)

def select_folder():
    global folder_path
    folder_path = filedialog.askdirectory()
    print("Selected table files path:", folder_path)

def output_folder():
    global output_folder
    print("Selected output folder path is:", output_folder)

def fixAmount(row=''):
    new_row = str(row).replace(',', '').replace('$', '').replace('USD', '').replace('US', '')
    try:
        new_row = float(new_row)
    except:
        new_row = float("0")
    return new_row

def process():
    print("\nProcess started")
    file_extension = os.path.splitext(excel_path)[1]
    if file_extension.lower() == ".csv":
        df = pd.read_csv(excel_path)
    elif file_extension.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(excel_path)
    else:
        print("Invalid file format. Only CSV, XLSX, and XLS formats are supported.")
    if "invoice_number" in df.columns and "total" in df.columns:
        df['total'] = df['total'].str.replace('US', '').str.replace('$', '').str.replace(',', '')
        df["Shipment Date"] = pd.to_datetime(df["Shipment Date"], errors='coerce')
        df["invoice_date"] = pd.to_datetime(df["invoice_date"], errors='coerce')
        df["Shipment Date"] = df["Shipment Date"].dt.strftime("%d/%m/%Y")
        df["invoice_date"] = df["invoice_date"].dt.strftime("%d/%m/%Y")
        df["Port of discharge"] = df["Port of discharge"].str.upper()

    column_names = df["base document"].unique().tolist()
    final_df = pd.DataFrame(columns=column_names)
    details_column = df.columns[df.columns != "base document"]
    final_df.insert(0, "Details", details_column)

    # Iterate over each row in the final_df DataFrame
    for index, row in final_df.iterrows():
        details_value = row['Details']
        if details_value in df.columns:
            values = df[details_value].values
            details_col_index = final_df.columns.get_loc('Details')
            for i, value in enumerate(values):
                if pd.isnull(value):
                    final_df.iloc[index, details_col_index + i + 1] = value
                elif all(value == v for v in values):
                    final_df.iloc[index, details_col_index + i + 1] = value
                else:
                    final_df.iloc[index, details_col_index + i + 1] = value

    # Save the final_df DataFrame to an Excel file
    document_types = ['PO', 'BL']
    matching_rows = df[df['base document'].isin(document_types)]
    matching_po_numbers = matching_rows.loc[matching_rows['base document'].isin(document_types), 'po_number']

    if not matching_po_numbers.empty:
        po_number = matching_po_numbers.iloc[0]
    else:
        po_number = None
    output_file_path = os.path.join(output_folder, f"{po_number}.xlsx")

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    # Write the DataFrame to the worksheet
    for row in dataframe_to_rows(final_df, index=False, header=True):
        ws.append(row)
    # Create a table
    table_range = f"A1:{chr(ord('A') + len(final_df.columns) - 1)}{len(final_df) + 1}"
    table = Table(displayName="Table1", ref=table_range)
    style = "TableStyleMedium9"  # You can choose a different table style if desired
    table.tableStyleInfo = TableStyleInfo(name=style)
    ws.add_table(table)
    # Sheet name
    ws.title = "Label data"
    wb.save(output_file_path)
    details_col_index = 1

    for row in range(2, ws.max_row + 1):
        details_cell = ws.cell(row=row, column=details_col_index)
        pi_cell = ws.cell(row=row, column=details_col_index + 1)
        po_cell = ws.cell(row=row, column=details_col_index + 2)
        ci_cell = ws.cell(row=row, column=details_col_index + 3)
        details_value = details_cell.value
        pi_value = pi_cell.value
        po_value = po_cell.value
        ci_value = ci_cell.value

        if any(pd.isnull(value) for value in (pi_value, po_value, ci_value)):
            # details_cell.font = Font(color="FFCC00")                  # Yellow font color
            details_cell.font = Font(color="0000FF")                    # Blue font color

        elif all(value == pi_value for value in (po_value, ci_value)):
            details_cell.font = Font(color="006400")                    # Green font color
        else:
            details_cell.font = Font(color="8B4513")                  # Brown font color
            # details_cell.font = Font(color="FF0000")                    # RED font color

    # Save the modified workbook
    wb.save(output_file_path)
    print("label data added to output file.")

 # -----------------------------------------------------------------------
    print("\nProcessing table data")
    # time.sleep(2)
    dfs = []
    file_paths = []
    base_documents = []

    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        file_extension = os.path.splitext(file_path)[1]
        if file_extension.lower() == ".csv":
            df = pd.read_csv(file_path)
        elif file_extension.lower() in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path)
        else:
            print("Invalid file format (Only CSV, XLSX, & XLS formats are supported).")
            continue
        # base document value
        base_document = df['base document'].iloc[0]
        # print("base doc...", base_document)
        base_documents.append(base_document)
        # Required Col
        df = df[['base document', 'Item Code', 'description', 'Quantity', 'unit_price', 'total_price']]

        # Converting unitprice column to string
        # df['unit_price'] = df['unit_price'].astype(str)
        # df['unit_price'] = df['unit_price'].str.replace('US', '').str.replace('$', '').str.replace(',', '').str.replace('.00', '')
        # df['unit_price'] = pd.to_numeric(df['unit_price'], errors='coerce')

        dfs.append(df)
        file_paths.append(file_path)
        if not dfs:
            print("No valid files found in the selected folder.")

        final_df = pd.concat(dfs, axis=0)
        final_df.sort_values('Item Code', inplace=True)
        # final_df['total_price'] = pd.to_numeric(final_df['total_price'].str.replace('US', '').str.replace('$', '').str.replace(',', '').str.replace('.00', ''), errors='coerce')

        final_df['total_price'] = final_df['total_price'].apply(lambda x:fixAmount(x))
        final_df['unit_price'] = final_df['unit_price'].apply(lambda x:fixAmount(x))


        final_df['Item No'] = final_df['Item Code']
        final_df['Description'] = final_df['description']
        final_df = final_df.drop(['Item Code', 'description'], axis=1)

        # groupby Item code
        final_df['Counter'] = final_df.groupby('Item No').cumcount() + 1
        final_df = final_df.pivot(index='Item No', columns='Counter', values=['Quantity', 'unit_price', 'total_price'])
        columns = []
        for col in final_df.columns:
            col_name = col[0]
            col_index = col[1]
            if col_index > 1:
                new_col_name = f'{col_name}_{col_index}'
            else:
                new_col_name = col_name + "_1"
            columns.append(new_col_name)
        final_df.columns = columns
        final_df = final_df.reset_index()

    # Calculation of Quantity, Unit Price & Total Price
    final_df['Quantity'] = final_df['Quantity_2'] - final_df['Quantity_1']
    final_df['Unit Price'] = final_df['unit_price_2']
    final_df['Total Price'] = final_df['total_price_2'] - final_df['total_price_1']

    total_row = final_df.sum(numeric_only=True)
    total_row['Item No'] = 'Total'
    total_row['Quantity_1'] = final_df['Quantity_1'].sum()
    total_row['Quantity_2'] = final_df['Quantity_2'].sum()
    total_row['Quantity'] = final_df['Quantity'].sum()
    total_row['unit_price_1'] = final_df['unit_price_1'].sum()
    total_row['unit_price_2'] = final_df['unit_price_2'].sum()
    total_row['Unit Price'] = final_df['Unit Price'].sum()
    total_row['total_price_1'] = final_df['total_price_1'].sum()
    total_row['total_price_2'] = final_df['total_price_2'].sum()
    total_row['Total Price'] = final_df['Total Price'].sum()

    # Append the total row to the DataFrame
    final_df = pd.concat([final_df, total_row.to_frame().T], ignore_index=True)

    # Reordering the column
    new_col_order = ['Item No', 'Quantity_1', 'unit_price_1', 'total_price_1', 'Quantity_2', 'unit_price_2', 'total_price_2',
                     'Quantity', 'Unit Price', 'Total Price']
    final_df = final_df[new_col_order]

    # Styled dataframe to Excel
    temp_file_path = os.path.join(output_folder, "temp.xlsx")
    # print("styled_df col", styled_df.columns)
    # styled_df.to_excel(temp_file_path, sheet_name='Table Data', index=False, startrow=1)
    final_df.to_excel(temp_file_path, sheet_name='Table Data', index=False, startrow=1)

    # Load the existing workbook
    wb = load_workbook(output_file_path)

    new_sheet_name = 'Table Data'
    wb.create_sheet(title=new_sheet_name)
    temp_wb = load_workbook(temp_file_path)
    temp_ws = temp_wb.active
    new_ws = wb[new_sheet_name]
    for row in temp_ws.iter_rows():
        for cell in row:
            new_ws[cell.coordinate].value = cell.value
    # number of rows & col
    num_rows = final_df.shape[0]
    num_cols = final_df.shape[1]

    def apply_border_to_range(worksheet, range_str):
        border = Border(top=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"),
                        left=Side(style="thin"))
        for row in worksheet[range_str]:
            for cell in row:
                cell.border = border
    # Apply grid lines to the entire dataframe range, including the last row
    start_cell = new_ws.cell(row=1, column=1)
    end_cell = new_ws.cell(row=num_rows + 2, column=num_cols)
    data_range = f"{start_cell.coordinate}:{end_cell.coordinate}"
    apply_border_to_range(new_ws, data_range)

    # Define the column names to merge and their corresponding merge ranges
    columns_to_merge = {
        'Quantity_1': ['B1', 'D1'],
        'Quantity_2': ['E1', 'G1'],
        'Quantity': ['H1', 'J1']
    }
    # Get the base document value from the list
    base_document_1 = base_documents[0]
    base_document_2 = base_documents[1]
    # New sheet
    new_sheet = wb[new_sheet_name]

    # Merging cell in the 0th row
    for col, merge_ranges in columns_to_merge.items():
        start_cell = new_ws[merge_ranges[0]]  # Get the start cell of the merge range
        end_cell = new_ws[merge_ranges[1]]  # Get the end cell of the merge range
        merge_range = f"{start_cell.coordinate}:{end_cell.coordinate}"
        new_ws .merge_cells(merge_range)

        # Apply alignment to the merged cell
        merged_cell = new_ws [start_cell.coordinate]
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 'Quantity_1':
            # Add the first value of the base document to the merged cell
            merged_cell.value = f"{base_document_1}"
        elif col == 'Quantity_2':
            # Add the second value of the base document to the merged cell
            merged_cell.value = f"{base_document_2}"
        elif col == 'Quantity':
            # Add the string value "Variance" to the merged cell
            merged_cell.value = "Variance"
            font = Font(color="FF0000")  # Apply red font color to the merged cell
            merged_cell.font = font

    # Get the header row
    header_row = new_ws[2]

    # Set red font color for values of specified columns
    font = Font(color="FF0000")  # Red font color
    for row in new_ws.iter_rows(min_row=3):
        for cell in row:
            if cell.column_letter in ['H', 'I', 'J']:
                cell.font = font


    # # Bold the 0th row, 1st row, and last row
    font_bold = Font(bold=True)
    for row_num, row in enumerate(new_ws.iter_rows(), start=1):
        if row_num == 2 or row_num == new_ws.max_row:
            for cell in row:
                cell.font = font_bold

    # Set red font color for column headers and values of specified columns
    # font = Font(color="00FF0000")  # Red font color
    font = Font(color="FF0000", bold=True)  # Red font color
    for cell in header_row:
        if cell.value in ['Quantity', 'Unit Price', 'Total Price']:
            cell.font = font

    first_row = final_df.iloc[0]
    # Apply bold style to all cells in the first row while preserving their color
    for col_num, cell_value in enumerate(first_row, start=1):
        cell = new_ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color=cell.font.color)


    # Save the changes
    wb.save(output_file_path)
    os.remove(temp_file_path)
    print("table data added to output file.")
    print("All Process completed & file saved in output folder.")

def destroy_me():
    global window, treadLoop
    answer = askyesno(title='Mindful Automation Pvt Ltd', message='Are you sure you want to Quit ?')
    if (answer):
        try:
           treadLoop.cancel()
        except:
            pp = 0
        current_system_pid = os.getpid()
        ThisSystem = psutil.Process(current_system_pid)
        ThisSystem.terminate()
        window.destroy()

root = tk.Tk()
root.title("Mindful Automation Pvt Ltd")
root.geometry("400x290")
root['bg'] = 'white'
current_path = os.path.dirname(os.path.realpath(__file__))
root.wm_iconbitmap(f"{current_path}/icons/mindful_logo.ico")

label = tk.Label(root, text="PAN EMIRATES", width=50, height=3, fg="#03001C", font=('Arial', 10, 'bold'))
label.grid(column=1, row=1)

input = tk.Button(root, text="Select label Excel File", command=input_file, height=1, width=32)
input.grid(column=1, row=2, pady=10)
input = tk.Button(root, text="Select Table Excel Folder", command=select_folder, height=1, width=32)
input.grid(column=1, row=3, pady=10)
output = tk.Button(root, text="Select Output Folder", command=output_folder, height=1, width=32)
output.grid(column=1, row=4, pady=10)
process = tk.Button(root, text="Start Process", command=process, height=1, width=32)
process.grid(column=1, row=5, pady=10)
end = tk.Button(root, text="EXIT", command=destroy_me, height=1, width=32)
end.grid(column=1, row=6, pady=10)
root.mainloop()
